import os
from datetime import datetime
from typing import List, Dict, Any
from pathlib import Path

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

# Excel generation
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class ProjectExportGenerator:
    """Generate PDF and Excel cutting lists from a project's parts."""
    
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def _sanitize(self, name: str) -> str:
        if not name:
            return "Sans_Nom"
        clean = "".join([c for c in name if c.isalnum() or c in (' ', '-', '_')]).strip()
        clean = clean.replace(' ', '_')
        return clean or "Sans_Nom"
        
    def _sort_parts(self, parts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Sort parts by material type (alphabetical) then by area (width * height) descending."""
        return sorted(
            parts, 
            key=lambda p: (
                str(p.get('material_name', '')), 
                -(p.get('width', 0) * p.get('height', 0))
            )
        )

    def _get_export_dir(self, client_name: str, project_name: str) -> Path:
        safe_client = self._sanitize(client_name)
        safe_project = self._sanitize(project_name)
        # Target: Moteur/UserData/Clients/[Nom_du_Client]/[Nom_du_Projet]/Fiches_de_Debit/
        final_dir = Path(self.output_dir) / "Clients" / safe_client / safe_project / "Fiches_de_Debit"
        final_dir.mkdir(parents=True, exist_ok=True)
        return final_dir

    def generate_all(self, parts: List[Dict[str, Any]], project_name: str, client_name: str, formats: List[str]) -> Dict[str, str]:
        if not parts:
            return {}
            
        timestamp = datetime.now().strftime("%Y-%m-%d_%Hh%M")
        export_dir = self._get_export_dir(client_name, project_name)
        
        sorted_parts = self._sort_parts(parts)
        files = {}
        
        base_name = f"Fiche_Debit_{self._sanitize(project_name)}_{timestamp}"
        
        if "pdf" in formats:
            pdf_path = self.generate_pdf(sorted_parts, project_name, client_name, str(export_dir), base_name)
            if pdf_path:
                files["pdf"] = str(pdf_path)
                
        if "excel" in formats:
            excel_path = self.generate_excel(sorted_parts, project_name, client_name, str(export_dir), base_name)
            if excel_path:
                files["excel"] = str(excel_path)
                
        return files

    def generate_pdf(self, parts: List[Dict[str, Any]], project_name: str, client_name: str, export_dir: str, base_name: str) -> str:
        file_path = os.path.join(export_dir, f"{base_name}.pdf")
        
        doc = SimpleDocTemplate(
            file_path,
            pagesize=A4,
            rightMargin=15*mm,
            leftMargin=15*mm,
            topMargin=20*mm,
            bottomMargin=15*mm
        )
        
        styles = getSampleStyleSheet()
        elements = []
        
        # Header
        title = Paragraph(f"<b>Fiche de Débit - Projet : {project_name}</b>", styles['Title'])
        client_info = Paragraph(f"Client : {client_name}", styles['Normal'])
        date_info = Paragraph(f"Date : {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal'])
        parts_count = Paragraph(f"Total des pièces : {len(parts)}", styles['Normal'])
        
        elements.extend([title, Spacer(1, 5*mm), client_info, date_info, parts_count, Spacer(1, 10*mm)])
        
        # Table data
        data = [["Nom", "Dim. (L x l)", "Épais.", "Matériau", "Qté", "Fait"]]
        
        for part in parts:
            name = part.get('name', 'Inconnu')
            width = part.get('width', 0)
            height = part.get('height', 0)
            thickness = part.get('material_thickness', '-')
            material = str(part.get('material_name', 'Inconnu'))[:25]
            quantity = part.get('quantity', 1)
            
            data.append([
                name,
                f"{width} x {height}",
                str(thickness),
                material,
                str(quantity),
                "" # Empty for checkbox
            ])
            
        # Table styling
        # A4 width is 210mm. Margins are 15+15=30mm. Available = 180mm.
        col_widths = [50*mm, 35*mm, 15*mm, 50*mm, 15*mm, 15*mm]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')), # Indigo-600
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (2, -1), 'CENTER'), # Center dimensions and thickness
            ('ALIGN', (4, 0), (-1, -1), 'CENTER'), # Center Qty and Fait
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            # Draw empty checkbox square
            ('BOX', (5, 1), (-1, -1), 0.5, colors.grey),
        ]))
        
        # Add alternating row colors
        for i in range(1, len(data)):
            if i % 2 == 0:
                table.setStyle(TableStyle([('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC'))]))
        
        elements.append(table)
        doc.build(elements)
        
        return file_path

    def generate_excel(self, parts: List[Dict[str, Any]], project_name: str, client_name: str, export_dir: str, base_name: str) -> str:
        file_path = os.path.join(export_dir, f"{base_name}.xlsx")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fiche_de_Debit"
        
        # Header Info
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Fiche de Débit - Projet : {project_name}"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "Client :"
        ws['B2'] = client_name
        ws['A3'] = "Date :"
        ws['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')
        ws['A4'] = "Total pièces :"
        ws['B4'] = len(parts)
        
        for row in range(2, 5):
            ws[f'A{row}'].font = Font(bold=True)
            
        # Table Headers
        headers = ["Nom de la Pièce", "Dimensions (L x l)", "Épaisseur", "Matériau", "Quantité", "Fait"]
        header_row = 6
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        ws.row_dimensions[header_row].height = 20
        
        # Data
        for row_idx, part in enumerate(parts, header_row + 1):
            name = part.get('name', 'Inconnu')
            width = part.get('width', 0)
            height = part.get('height', 0)
            thickness = part.get('material_thickness', '-')
            material = str(part.get('material_name', 'Inconnu'))
            quantity = part.get('quantity', 1)
            
            row_data = [
                name,
                f"{width} x {height}",
                thickness,
                material,
                quantity,
                ""  # Empty for checkbox
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = thin_border
                
                if col_idx in [2, 3, 5, 6]: # Center dimensions, thickness, qty, checkbox
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
        # Calculate optimal column widths
        col_letters = ['A', 'B', 'C', 'D', 'E', 'F']
        for idx, col_letter in enumerate(col_letters):
            max_length = 0
            for cell in ws[col_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            # Custom adjustments
            if col_letter == 'A': adjusted_width = max(25, adjusted_width)
            if col_letter == 'C': adjusted_width = max(12, adjusted_width)
            if col_letter == 'D': adjusted_width = max(20, adjusted_width)
            if col_letter == 'E': adjusted_width = max(10, adjusted_width)
            if col_letter == 'F': adjusted_width = max(8, adjusted_width)
            ws.column_dimensions[col_letter].width = adjusted_width
            
        wb.save(file_path)
        return file_path
